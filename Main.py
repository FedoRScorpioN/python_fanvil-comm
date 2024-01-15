import logging
import os
import subprocess
import requests
from requests.auth import HTTPDigestAuth

import scapy.all as sc
from openpyxl import load_workbook

import DB
from DB import create_tables
from dotenv import load_dotenv
import json

from flask import Flask, request

app = Flask(__name__)

load_dotenv()
SUBNET_1 = os.getenv("SUBNET_1")
SUBNET_2 = os.getenv("SUBNET_2")
EXCEL_FILE = os.getenv("EXCEL_FILE")
JSON_FILE = os.getenv("JSON_FILE")
HTML_FILE = os.getenv("HTML_FILE")

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
handler = logging.FileHandler('app.log', encoding='utf-8')
handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logging.getLogger().addHandler(handler)


def clear_mac_ip_pairs_table():
    try:
        if DB.conn is None or DB.cursor is None:
            logging.error('Отсутствует соединение с базой данных или указатель.')
            return

        DB.cursor.execute("DELETE FROM mac_ip_pairs;")
        DB.conn.commit()

        logging.info("Таблица mac_ip_pairs успешно очищена.")
    except Exception as e:
        logging.error(f'Ошибка при очистке таблицы mac_ip_pairs: {str(e)}')


def get_ip_mac_network(subnet):
    try:
        logging.info(f'Начинаем сканирование сети {subnet}')
        answered_list = sc.srp(sc.Ether(dst='ff:ff:ff:ff:ff:ff') / sc.ARP(pdst=subnet), timeout=1, verbose=False)[0]
        clients_list = [{'ip': element[1].psrc, 'mac': element[1].hwsrc} for element in answered_list]
        logging.info(f'Получены данные по MAC-адресам и IP-адресам: {clients_list}')
        return clients_list
    except Exception as e:
        logging.error(f'Ошибка в get_ip_mac_network: {str(e)}')
        return []


def ip_mac_to_database(mac_ip_list):
    try:
        if DB.conn is None or DB.cursor is None:
            logging.error('Отсутствует соединение с базой данных или указатель.')
            return

        for client in mac_ip_list:
            mac_address_upper = client['mac'].upper()

            DB.cursor.execute('''
                INSERT INTO mac_ip_pairs (mac_address, ip_address)
                VALUES (%s, %s)
            ''', (mac_address_upper, client['ip']))

        DB.conn.commit()
        logging.info('Данные добавлены в базу данных.')
    except Exception as e:
        logging.error(f'Ошибка при добавлении данных в базу данных: {str(e)}')


def scan_and_to_database(subnet):
    ip_mac_network = get_ip_mac_network(subnet)
    ip_mac_network.sort(key=lambda x: int(x['ip'].split('.')[3]))
    ip_mac_to_database(ip_mac_network)


def convert_excel_to_json(excel_file_path):
    try:
        workbook = load_workbook(excel_file_path)
        json_data = []

        for sheet in workbook.sheetnames:
            for row in workbook[sheet].iter_rows(min_row=4, values_only=True):
                json_row = dict(zip(get_column_names(), row))
                json_data.append(json_row)

                logging.info(f'Прочитана строка из Excel: {json_row}')

        with open(JSON_FILE, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, ensure_ascii=False, indent=2)

        return JSON_FILE
    except Exception as e:
        logging.error(f'Ошибка при конвертации Excel в JSON: {str(e)}')
        return None


def update_html_with_json(json_file_path, html_file_path):
    try:
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            json_data = json.load(json_file)

        json_string = json.dumps(json_data, ensure_ascii=False, indent=2)

        with open(html_file_path, 'r', encoding='utf-8') as html_file:
            html_content = html_file.read()

        start_index = html_content.find("var jsonData =")
        end_index = html_content.find(";", start_index)

        if start_index != -1 and end_index != -1:
            html_content = html_content[:start_index + len("var jsonData =\n")] + json_string + html_content[end_index:]

            logging.info(f'Обновлен HTML с использованием следующих данных JSON:\n{{ {json_string} }}')

            with open(html_file_path, 'w', encoding='utf-8') as html_file:
                html_file.write(html_content)

            logging.info('HTML успешно обновлен.')
        else:
            logging.error('Не удалось найти блок данных JSON в HTML файле.')

    except Exception as e:
        logging.error(f'Ошибка при обновлении HTML с использованием JSON: {str(e)}')


def get_column_names():
    return ["Должность", "ФИО", "Рабочий", "Внутренний", "Сотовый", "Местоположение", "Email", "Дата рождения"]


if __name__ == "__main__":
    create_tables()
    clear_mac_ip_pairs_table()

    scan_and_to_database(SUBNET_1)
    scan_and_to_database(SUBNET_2)

    JSON_FILE = convert_excel_to_json(EXCEL_FILE)

    if JSON_FILE:
        update_html_with_json(JSON_FILE, HTML_FILE)

    logging.info('Программа успешно выполнена.')
